from __future__ import annotations

import csv
from pathlib import Path

from bs4 import BeautifulSoup
from docx import Document as DocxDocument
from openpyxl import load_workbook
from pypdf import PdfReader

from client_profiler.models import DocumentInput


class UnsupportedFileTypeError(ValueError):
    pass


class DocumentReader:
    SUPPORTED_SUFFIXES = {
        ".txt",
        ".md",
        ".markdown",
        ".html",
        ".htm",
        ".pdf",
        ".docx",
        ".xlsx",
        ".csv",
    }

    def read(self, path: Path) -> DocumentInput:
        suffix = path.suffix.lower()
        if suffix not in self.SUPPORTED_SUFFIXES:
            raise UnsupportedFileTypeError(f"Unsupported file type: {suffix}")

        if suffix in {".txt", ".md", ".markdown"}:
            text = path.read_text(encoding="utf-8", errors="ignore")
            source_type = "text"
        elif suffix in {".html", ".htm"}:
            text = self._read_html(path)
            source_type = "html"
        elif suffix == ".pdf":
            text = self._read_pdf(path)
            source_type = "pdf"
        elif suffix == ".docx":
            text = self._read_docx(path)
            source_type = "docx"
        elif suffix == ".xlsx":
            text = self._read_xlsx(path)
            source_type = "xlsx"
        elif suffix == ".csv":
            text = self._read_csv(path)
            source_type = "csv"
        else:
            raise UnsupportedFileTypeError(f"Unsupported file type: {suffix}")

        return DocumentInput(
            source_path=path,
            source_type=source_type,
            text=text,
            metadata={"suffix": suffix, "filename": path.name},
        )

    def _read_html(self, path: Path) -> str:
        html = path.read_text(encoding="utf-8", errors="ignore")
        soup = BeautifulSoup(html, "html.parser")
        return soup.get_text("\n", strip=True)

    def _read_pdf(self, path: Path) -> str:
        reader = PdfReader(str(path))
        pages = []
        for page in reader.pages:
            pages.append(page.extract_text() or "")
        return "\n\n".join(pages)

    def _read_docx(self, path: Path) -> str:
        doc = DocxDocument(str(path))
        return "\n".join(p.text for p in doc.paragraphs if p.text)

    def _read_xlsx(self, path: Path) -> str:
        wb = load_workbook(filename=str(path), data_only=True)
        lines: list[str] = []
        for ws in wb.worksheets:
            lines.append(f"[Sheet: {ws.title}]")
            for row in ws.iter_rows(values_only=True):
                row_values = ["" if cell is None else str(cell) for cell in row]
                if any(v.strip() for v in row_values):
                    lines.append(" | ".join(row_values))
        return "\n".join(lines)

    def _read_csv(self, path: Path) -> str:
        rows: list[str] = []
        with path.open("r", encoding="utf-8", errors="ignore", newline="") as f:
            reader = csv.reader(f)
            for row in reader:
                rows.append(" | ".join(row))
        return "\n".join(rows)
